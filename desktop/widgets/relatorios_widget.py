from datetime import date, datetime
from html import escape
import os
import re
import unicodedata

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QColor, QFont
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QDateEdit,
    QFileDialog,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QProgressBar,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from access_control import get_action_label, has_action_access
from api_client import api_client
from user_preferences import (
    apply_table_column_widths,
    get_table_column_widths,
    get_widget_preferences,
    save_widget_preferences,
)
from widgets.filter_utils import filter_value, is_all_option
from widgets.table_utils import configure_data_table, refresh_data_table_layout


CALENDAR_STYLE = """
    QDateEdit {
        background-color: #ffffff;
        color: #1e293b;
        border: 1px solid #cbd5e1;
        border-radius: 8px;
        padding: 8px 12px;
        min-height: 32px;
        font-size: 13px;
    }
    QDateEdit::drop-down {
        border: none;
        width: 24px;
    }
    QDateEdit::down-arrow {
        image: none;
    }
"""


class RelatoriosWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.usuario = {}
        self._restoring_tab_state = False
        self._loaded_tabs = {
            "movimentacoes": False,
            "estoque": False,
            "pedidos": False,
            "demandas": False,
        }
        self._report_sources = {
            "movimentacoes": [],
            "estoque": [],
            "pedidos": [],
            "demandas": [],
        }
        self._report_visible = {
            "movimentacoes": [],
            "estoque": [],
            "pedidos": [],
            "demandas": [],
        }
        self._summary_labels = {}
        self._result_labels = {}
        self._detail_labels = {}
        self._detail_empty_messages = {}
        self._empty_state_cards = {}
        self.init_ui()

    def set_usuario(self, usuario):
        self.usuario = usuario or {}
        self.aplicar_permissoes()
        self._restore_saved_tab()

    def _pode(self, action_key):
        return has_action_access(self.usuario, action_key)

    def _avisar_sem_permissao(self, action_key):
        QMessageBox.warning(
            self,
            "Acesso não permitido",
            f"Você não tem permissão para {get_action_label(action_key)}.",
        )

    def on_show(self):
        pass

    def showEvent(self, event):
        self._apply_theme_styles()
        super().showEvent(event)

    def _is_dark_theme(self):
        app = QApplication.instance()
        return str(app.property("accessibility_theme") or "Claro") == "Escuro"

    def _theme_colors(self):
        if self._is_dark_theme():
            return {
                "hero_bg": "rgba(15, 23, 42, 0.55)",
                "hero_border": "rgba(96, 165, 250, 0.18)",
                "card_bg": "rgba(15, 23, 42, 0.34)",
                "card_border": "rgba(148, 163, 184, 0.16)",
                "title": "#f8fafc",
                "muted": "#94a3b8",
            }
        return {
            "hero_bg": "rgba(255, 255, 255, 0.98)",
            "hero_border": "rgba(148, 163, 184, 0.18)",
            "card_bg": "rgba(248, 250, 252, 1.0)",
            "card_border": "rgba(203, 213, 225, 0.9)",
            "title": "#0f172a",
            "muted": "#64748b",
        }

    def _apply_theme_styles(self):
        colors_map = self._theme_colors()
        self.setStyleSheet(
            f"""
            QFrame#reportHeroCard {{
                background-color: {colors_map['hero_bg']};
                border: 1px solid {colors_map['hero_border']};
                border-radius: 20px;
            }}
            QLabel#reportHeroTitle {{
                color: {colors_map['title']};
                font-size: 22px;
                font-weight: 700;
            }}
            QLabel#reportHeroSubtitle {{
                color: {colors_map['muted']};
                font-size: 13px;
            }}
            QFrame#reportSummaryCard, QFrame#reportActionsCard {{
                background-color: {colors_map['card_bg']};
                border: 1px solid {colors_map['card_border']};
                border-radius: 16px;
            }}
            QLabel#reportSummaryTitle {{
                color: {colors_map['muted']};
                font-size: 11px;
                font-weight: 700;
            }}
            QLabel#reportSummaryValue {{
                color: {colors_map['title']};
                font-size: 24px;
                font-weight: 700;
            }}
            QLabel#reportSummaryCaption {{
                color: {colors_map['muted']};
                font-size: 12px;
            }}
            QLabel#reportDetailTitle {{
                color: {colors_map['title']};
                font-size: 16px;
                font-weight: 700;
            }}
            QLabel#reportDetailBody {{
                color: {colors_map['muted']};
                font-size: 12px;
            }}
            QLabel#reportResultsLabel {{
                color: {colors_map['muted']};
                font-size: 12px;
                font-weight: 600;
                padding-left: 2px;
            }}
            QFrame#reportEmptyStateCard {{
                background-color: {colors_map['card_bg']};
                border: 1px dashed {colors_map['card_border']};
                border-radius: 16px;
            }}
            QLabel#reportEmptyStateTitle {{
                color: {colors_map['title']};
                font-size: 15px;
                font-weight: 700;
            }}
            QLabel#reportEmptyStateBody {{
                color: {colors_map['muted']};
                font-size: 12px;
            }}
            QLineEdit#reportSearchInput {{
                background-color: {colors_map['card_bg']};
                color: {colors_map['title']};
                border: 1px solid {colors_map['card_border']};
                border-radius: 12px;
                padding: 10px 14px;
                min-height: 18px;
            }}
            """
        )

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(20)

        header_card = QFrame()
        header_card.setObjectName("reportHeroCard")
        header_layout = QVBoxLayout(header_card)
        header_layout.setContentsMargins(22, 20, 22, 20)
        header_layout.setSpacing(8)

        title = QLabel("Relatórios do Sistema")
        title.setObjectName("reportHeroTitle")
        header_layout.addWidget(title)

        subtitle = QLabel(
            "Acompanhe movimentacoes, estoque, pedidos e demandas com filtros, resumos visuais e exportacao."
        )
        subtitle.setObjectName("reportHeroSubtitle")
        subtitle.setWordWrap(True)
        header_layout.addWidget(subtitle)
        layout.addWidget(header_card)

        self.tabs = QTabWidget()
        self.tabs.setObjectName("paramTabs")
        self.tabs.currentChanged.connect(self.on_tab_changed)

        self.tab_movimentacoes = self.create_tab_movimentacoes()
        self.tabs.addTab(self.tab_movimentacoes, "Movimentações")

        self.tab_estoque = self.create_tab_estoque()
        self.tabs.addTab(self.tab_estoque, "Estoque")

        self.tab_pedidos = self.create_tab_pedidos()
        self.tabs.addTab(self.tab_pedidos, "Pedidos")

        self.tab_demandas = self.create_tab_demandas()
        self.tabs.addTab(self.tab_demandas, "Demandas")

        layout.addWidget(self.tabs)
        self.aplicar_permissoes()

    def _restore_saved_tab(self):
        preferences = get_widget_preferences(self.usuario, "relatorios")
        active_tab = preferences.get("active_tab")
        if not isinstance(active_tab, str):
            return

        for index in range(self.tabs.count()):
            if self.tabs.tabText(index) == active_tab:
                self._restoring_tab_state = True
                try:
                    self.tabs.setCurrentIndex(index)
                finally:
                    self._restoring_tab_state = False
                return

    def _save_active_tab(self):
        current_index = self.tabs.currentIndex()
        if current_index < 0:
            return

        preferences = get_widget_preferences(self.usuario, "relatorios")
        preferences["active_tab"] = self.tabs.tabText(current_index)
        save_widget_preferences(self.usuario, "relatorios", preferences)

    def _widths_pref_key(self, tipo):
        return f"{tipo}_widths"

    def _save_table_widths(self, tipo):
        if not self.usuario:
            return

        table = self._table_for_tipo(tipo)
        if table is None:
            return

        preferences = get_widget_preferences(self.usuario, "relatorios")
        preferences[self._widths_pref_key(tipo)] = get_table_column_widths(table)
        save_widget_preferences(self.usuario, "relatorios", preferences)

    def _apply_saved_table_widths(self, tipo):
        table = self._table_for_tipo(tipo)
        if table is None:
            return

        preferences = get_widget_preferences(self.usuario, "relatorios")
        apply_table_column_widths(table, preferences.get(self._widths_pref_key(tipo)))

    def aplicar_permissoes(self):
        pode_exportar = self._pode("relatorios.export")
        for attr_name in (
            "btn_exportar_excel",
            "btn_exportar_pdf",
            "btn_exportar_excel_estoque",
            "btn_exportar_pdf_estoque",
            "btn_exportar_excel_pedidos",
            "btn_exportar_pdf_pedidos",
            "btn_exportar_excel_demandas",
            "btn_exportar_pdf_demandas",
        ):
            btn = getattr(self, attr_name, None)
            if btn is not None:
                btn.setVisible(pode_exportar)

    def on_tab_changed(self, index):
        if not self._restoring_tab_state:
            self._save_active_tab()
        tab_text = self.tabs.tabText(index)
        if tab_text == "Movimentações" and not self._loaded_tabs["movimentacoes"]:
            self.carregar_movimentacoes()
            self.carregar_empresas_movimentacoes()
            self._loaded_tabs["movimentacoes"] = True
        elif tab_text == "Estoque" and not self._loaded_tabs["estoque"]:
            self.carregar_categorias()
            self.carregar_estoque()
            self._loaded_tabs["estoque"] = True
        elif tab_text == "Pedidos" and not self._loaded_tabs["pedidos"]:
            self.carregar_pedidos()
            self.carregar_empresas_pedidos()
            self._loaded_tabs["pedidos"] = True
        elif tab_text == "Demandas" and not self._loaded_tabs["demandas"]:
            self.carregar_demandas()
            self.carregar_empresas_demandas()
            self._loaded_tabs["demandas"] = True

    def _create_summary_strip(self, key, specs):
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)
        self._summary_labels[key] = {}

        for summary_key, title, caption in specs:
            card = QFrame()
            card.setObjectName("reportSummaryCard")
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(16, 14, 16, 14)
            card_layout.setSpacing(4)

            title_label = QLabel(title)
            title_label.setObjectName("reportSummaryTitle")
            card_layout.addWidget(title_label)

            value_label = QLabel("--")
            value_label.setObjectName("reportSummaryValue")
            card_layout.addWidget(value_label)

            caption_label = QLabel(caption)
            caption_label.setObjectName("reportSummaryCaption")
            caption_label.setWordWrap(True)
            card_layout.addWidget(caption_label)

            layout.addWidget(card, 1)
            self._summary_labels[key][summary_key] = value_label

        return container

    def _create_actions_card(self, primary_button, export_buttons):
        card = QFrame()
        card.setObjectName("reportActionsCard")
        layout = QHBoxLayout(card)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)
        layout.addWidget(primary_button)
        layout.addStretch()
        for button in export_buttons:
            layout.addWidget(button)
        return card

    def _set_summary_value(self, group, key, value):
        label = self._summary_labels.get(group, {}).get(key)
        if label is not None:
            label.setText(str(value))

    def _create_search_card(self, placeholder, callback):
        card = QFrame()
        card.setObjectName("reportActionsCard")
        layout = QHBoxLayout(card)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)

        title = QLabel("Busca rapida")
        title.setObjectName("reportSummaryTitle")
        layout.addWidget(title)

        search_input = QLineEdit()
        search_input.setObjectName("reportSearchInput")
        search_input.setPlaceholderText(placeholder)
        search_input.setClearButtonEnabled(True)
        search_input.textChanged.connect(callback)
        layout.addWidget(search_input, 1)
        return card, search_input

    def _create_detail_panel(self, key, title, empty_message):
        card = QFrame()
        card.setObjectName("reportActionsCard")
        card.setMinimumWidth(320)
        card.setMaximumWidth(380)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(10)

        title_label = QLabel(title)
        title_label.setObjectName("reportDetailTitle")
        layout.addWidget(title_label)

        body = QLabel(empty_message)
        body.setObjectName("reportDetailBody")
        body.setWordWrap(True)
        body.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        body.setTextFormat(Qt.RichText)
        body.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(body, 1)
        layout.addStretch()

        self._detail_labels[key] = body
        self._detail_empty_messages[key] = empty_message
        return card

    def _create_table_panel(self, key, table, empty_title, empty_message):
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        result_label = QLabel("Aguardando dados do relatorio...")
        result_label.setObjectName("reportResultsLabel")
        layout.addWidget(result_label)

        layout.addWidget(table, 1)

        empty_card = QFrame()
        empty_card.setObjectName("reportEmptyStateCard")
        empty_layout = QVBoxLayout(empty_card)
        empty_layout.setContentsMargins(22, 20, 22, 20)
        empty_layout.setSpacing(8)

        title_label = QLabel(empty_title)
        title_label.setObjectName("reportEmptyStateTitle")
        empty_layout.addWidget(title_label)

        body_label = QLabel(empty_message)
        body_label.setObjectName("reportEmptyStateBody")
        body_label.setWordWrap(True)
        empty_layout.addWidget(body_label)
        empty_layout.addStretch()

        empty_card.hide()
        layout.addWidget(empty_card, 1)

        self._result_labels[key] = result_label
        self._empty_state_cards[key] = empty_card
        return container

    def _set_detail_empty(self, key):
        label = self._detail_labels.get(key)
        if label is not None:
            label.setText(self._detail_empty_messages.get(key, "Selecione um registro para ver os detalhes."))

    def _set_report_feedback(self, key, count):
        label = self._result_labels.get(key)
        if label is None:
            return

        descriptors = {
            "movimentacoes": ("movimentacao encontrada", "movimentacoes encontradas"),
            "estoque": ("item encontrado", "itens encontrados"),
            "pedidos": ("pedido encontrado", "pedidos encontrados"),
            "demandas": ("demanda encontrada", "demandas encontradas"),
        }
        singular, plural = descriptors.get(key, ("registro encontrado", "registros encontrados"))
        if count == 0:
            label.setText("Nenhum resultado com os filtros atuais.")
        elif count == 1:
            label.setText(f"1 {singular}.")
        else:
            label.setText(f"{count} {plural}.")

    def _set_table_empty_state(self, key, has_rows):
        table = self._table_for_tipo(key)
        empty_card = self._empty_state_cards.get(key)
        if table is not None:
            table.setVisible(has_rows)
        if empty_card is not None:
            empty_card.setVisible(not has_rows)

    def _build_detail_html(self, title, fields):
        colors_map = self._theme_colors()
        parts = [
            f"<div style='font-size:16px;font-weight:700;color:{colors_map['title']}; margin-bottom:8px;'>{escape(str(title))}</div>"
        ]
        for label, value in fields:
            text = "-" if value in (None, "", "None") else str(value)
            text = escape(text).replace("\n", "<br>")
            parts.append(
                f"<div style='margin-bottom:6px;'><span style='font-weight:700;color:{colors_map['title']};'>{escape(str(label))}:</span> "
                f"<span style='color:{colors_map['muted']};'>{text}</span></div>"
            )
        return "".join(parts)

    def _apply_text_search(self, records, query, fields):
        term = str(query or "").strip().lower()
        if not term:
            return list(records)

        filtered = []
        for record in records:
            haystack = " ".join(str(record.get(field, "") or "") for field in fields).lower()
            if term in haystack:
                filtered.append(record)
        return filtered

    def _find_visible_record(self, key, record_id):
        for record in self._report_visible.get(key, []):
            if str(record.get("id", "")) == str(record_id):
                return record
        return None

    def _update_detail_from_selection(self, key):
        table = self._table_for_tipo(key)
        if table is None or table.rowCount() == 0:
            self._set_detail_empty(key)
            return

        selected_items = table.selectedItems()
        if not selected_items:
            self._set_detail_empty(key)
            return

        row = selected_items[0].row()
        id_item = table.item(row, 0)
        if id_item is None:
            self._set_detail_empty(key)
            return

        record = self._find_visible_record(key, id_item.text())
        if not record:
            self._set_detail_empty(key)
            return

        if key == "movimentacoes":
            html = self._build_detail_html(
                f"Movimentacao #{record.get('id', '-')}",
                [
                    ("Material", record.get("material_nome")),
                    ("Tipo", str(record.get("tipo", "")).upper()),
                    ("Quantidade", record.get("quantidade")),
                    ("Empresa", record.get("empresa")),
                    ("Destinatario", record.get("destinatario")),
                    ("Usuário", record.get("usuario_nome")),
                    ("Data/Hora", record.get("data_hora")),
                    ("Observação", record.get("observacao")),
                ],
            )
        elif key == "estoque":
            html = self._build_detail_html(
                f"Material #{record.get('id', '-')}",
                [
                    ("Nome", record.get("nome")),
                    ("Descrição", record.get("descricao")),
                    ("Quantidade", record.get("quantidade")),
                    ("Categoria", record.get("categoria")),
                    ("Empresa", record.get("empresa")),
                    ("Status", str(record.get("status", "")).upper()),
                ],
            )
        elif key == "pedidos":
            html = self._build_detail_html(
                f"Pedido #{record.get('id', '-')}",
                [
                    ("Material", record.get("material_nome")),
                    ("Quantidade", record.get("quantidade")),
                    ("Solicitante", record.get("solicitante")),
                    ("Empresa", record.get("empresa")),
                    ("Departamento", record.get("departamento")),
                    ("Status", str(record.get("status", "")).upper()),
                    ("Data solicitacao", record.get("data_solicitacao")),
                    ("Data conclusão", record.get("data_conclusao") or "-"),
                    ("Observação", record.get("observacao")),
                    ("Link", record.get("link_compra")),
                ],
            )
        else:
            html = self._build_detail_html(
                f"Demanda #{record.get('id', '-')}",
                [
                    ("Título", record.get("titulo")),
                    ("Descrição", record.get("descricao")),
                    ("Solicitante", record.get("solicitante")),
                    ("Empresa", record.get("empresa")),
                    ("Prioridade", str(record.get("prioridade", "")).upper()),
                    ("Status", str(record.get("status", "")).upper()),
                    ("Data abertura", record.get("data_abertura")),
                    ("Responsável", record.get("responsavel")),
                    ("Observação", record.get("observacao")),
                ],
            )

        label = self._detail_labels.get(key)
        if label is not None:
            label.setText(html)

    def _build_table(self, headers, stretch_columns, minimum_widths):
        table = QTableWidget()
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.verticalHeader().setVisible(False)
        table.setSortingEnabled(True)
        table.setStyleSheet(
            """
            QTableWidget::item { padding: 10px 8px; }
            QHeaderView::section { padding: 10px 12px; }
            """
        )
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        configure_data_table(
            table,
            stretch_columns=stretch_columns,
            minimum_section_size=88,
            minimum_widths=minimum_widths,
        )
        return table

    def create_tab_movimentacoes(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(15)
        layout.addWidget(
            self._create_summary_strip(
                "movimentacoes",
                (
                    ("total", "Movimentações", "Registros encontrados"),
                    ("entradas", "Entradas", "Fluxo de entrada"),
                    ("saidas", "Saidas", "Fluxo de saida"),
                ),
            )
        )

        filtros = QGroupBox("Filtros")
        filtros.setObjectName("configGroup")
        filtros_layout = QHBoxLayout(filtros)
        filtros_layout.setContentsMargins(20, 15, 20, 15)
        filtros_layout.addWidget(QLabel("Data Inicio:"))
        self.mov_data_inicio = self._create_date_edit(QDate.currentDate().addMonths(-1))
        filtros_layout.addWidget(self.mov_data_inicio)
        filtros_layout.addWidget(QLabel("Data Fim:"))
        self.mov_data_fim = self._create_date_edit(QDate.currentDate())
        filtros_layout.addWidget(self.mov_data_fim)
        filtros_layout.addWidget(QLabel("Tipo:"))
        self.mov_tipo = QComboBox()
        self.mov_tipo.addItems(["Todos", "Entrada", "Saida"])
        self.mov_tipo.setObjectName("configCombo")
        filtros_layout.addWidget(self.mov_tipo)
        filtros_layout.addWidget(QLabel("Empresa:"))
        self.mov_empresa = QComboBox()
        self.mov_empresa.addItems(["Todas"])
        self.mov_empresa.setObjectName("configCombo")
        filtros_layout.addWidget(self.mov_empresa)
        filtros_layout.addStretch()
        layout.addWidget(filtros)

        mov_search_card, self.mov_search = self._create_search_card(
            "Pesquisar por material, usuário, empresa ou observação...",
            self._render_movimentacoes,
        )
        layout.addWidget(mov_search_card)

        self.btn_atualizar = QPushButton("Atualizar dados")
        self.btn_atualizar.setObjectName("btnPrimary")
        self.btn_atualizar.clicked.connect(self.atualizar_movimentacoes)
        self.btn_exportar_excel = QPushButton("Exportar Excel")
        self.btn_exportar_excel.setObjectName("btnSecondary")
        self.btn_exportar_excel.clicked.connect(lambda: self.exportar_excel("movimentacoes"))
        self.btn_exportar_pdf = QPushButton("Exportar PDF")
        self.btn_exportar_pdf.setObjectName("btnSecondary")
        self.btn_exportar_pdf.clicked.connect(lambda: self.exportar_pdf("movimentacoes"))
        layout.addWidget(self._create_actions_card(self.btn_atualizar, (self.btn_exportar_excel, self.btn_exportar_pdf)))

        self.mov_tabela = self._build_table(
            ["ID", "Material", "Tipo", "Quantidade", "Empresa", "Destinatario", "Data/Hora", "Usuário", "Observação"],
            stretch_columns=(1, 8),
            minimum_widths={0: 72, 1: 220, 2: 110, 3: 100, 4: 170, 5: 180, 6: 155, 7: 160, 8: 240},
        )
        self.mov_tabela.horizontalHeader().sectionResized.connect(
            lambda *_args, report_key="movimentacoes": self._save_table_widths(report_key)
        )
        self.mov_tabela.itemSelectionChanged.connect(lambda: self._update_detail_from_selection("movimentacoes"))
        mov_content = QHBoxLayout()
        mov_content.setContentsMargins(0, 0, 0, 0)
        mov_content.setSpacing(16)
        mov_content.addWidget(
            self._create_table_panel(
                "movimentacoes",
                self.mov_tabela,
                "Nenhuma movimentacao encontrada",
                "Ajuste a busca ou revise o periodo e os filtros para encontrar registros.",
            ),
            1,
        )
        mov_content.addWidget(
            self._create_detail_panel("movimentacoes", "Detalhes da movimentacao", "Selecione uma movimentacao para ver o resumo completo."),
            0,
        )
        layout.addLayout(mov_content)

        self.mov_progress = QProgressBar()
        self.mov_progress.setVisible(False)
        layout.addWidget(self.mov_progress)
        return widget

    def create_tab_estoque(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(15)
        layout.addWidget(
            self._create_summary_strip(
                "estoque",
                (
                    ("total", "Itens", "Materiais listados"),
                    ("ativos", "Ativos", "Disponiveis no estoque"),
                    ("criticos", "Críticos", "Quantidade baixa ou zerada"),
                ),
            )
        )

        filtros = QGroupBox("Filtros")
        filtros.setObjectName("configGroup")
        filtros_layout = QHBoxLayout(filtros)
        filtros_layout.setContentsMargins(20, 15, 20, 15)
        filtros_layout.addWidget(QLabel("Categoria:"))
        self.est_categoria = QComboBox()
        self.est_categoria.addItems(["Todas"])
        self.est_categoria.setObjectName("configCombo")
        filtros_layout.addWidget(self.est_categoria)
        filtros_layout.addWidget(QLabel("Empresa:"))
        self.est_empresa = QComboBox()
        self.est_empresa.addItems(["Todas"])
        self.est_empresa.setObjectName("configCombo")
        filtros_layout.addWidget(self.est_empresa)
        filtros_layout.addWidget(QLabel("Status:"))
        self.est_status = QComboBox()
        self.est_status.addItems(["Todos", "Ativo", "Inativo", "Descontinuado"])
        self.est_status.setObjectName("configCombo")
        filtros_layout.addWidget(self.est_status)
        filtros_layout.addStretch()
        layout.addWidget(filtros)

        est_search_card, self.est_search = self._create_search_card(
            "Pesquisar por nome, categoria, empresa ou status...",
            self._render_estoque,
        )
        layout.addWidget(est_search_card)

        self.btn_carregar_estoque = QPushButton("Atualizar dados")
        self.btn_carregar_estoque.setObjectName("btnPrimary")
        self.btn_carregar_estoque.clicked.connect(self.atualizar_estoque)
        self.btn_exportar_excel_estoque = QPushButton("Exportar Excel")
        self.btn_exportar_excel_estoque.setObjectName("btnSecondary")
        self.btn_exportar_excel_estoque.clicked.connect(lambda: self.exportar_excel("estoque"))
        self.btn_exportar_pdf_estoque = QPushButton("Exportar PDF")
        self.btn_exportar_pdf_estoque.setObjectName("btnSecondary")
        self.btn_exportar_pdf_estoque.clicked.connect(lambda: self.exportar_pdf("estoque"))
        layout.addWidget(
            self._create_actions_card(
                self.btn_carregar_estoque,
                (self.btn_exportar_excel_estoque, self.btn_exportar_pdf_estoque),
            )
        )

        self.est_tabela = self._build_table(
            ["ID", "Nome", "Descrição", "Quantidade", "Categoria", "Empresa", "Status"],
            stretch_columns=(1, 2),
            minimum_widths={0: 72, 1: 220, 2: 280, 3: 100, 4: 160, 5: 180, 6: 120},
        )
        self.est_tabela.horizontalHeader().sectionResized.connect(
            lambda *_args, report_key="estoque": self._save_table_widths(report_key)
        )
        self.est_tabela.itemSelectionChanged.connect(lambda: self._update_detail_from_selection("estoque"))
        est_content = QHBoxLayout()
        est_content.setContentsMargins(0, 0, 0, 0)
        est_content.setSpacing(16)
        est_content.addWidget(
            self._create_table_panel(
                "estoque",
                self.est_tabela,
                "Nenhum item encontrado",
                "Ajuste a busca ou revise categoria, empresa e status para localizar materiais.",
            ),
            1,
        )
        est_content.addWidget(
            self._create_detail_panel("estoque", "Detalhes do material", "Selecione um item do estoque para ver os detalhes."),
            0,
        )
        layout.addLayout(est_content)
        return widget

    def create_tab_pedidos(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(15)
        layout.addWidget(
            self._create_summary_strip(
                "pedidos",
                (
                    ("total", "Pedidos", "Solicitacoes no periodo"),
                    ("pendentes", "Pendentes", "Aguardando tratativa"),
                    ("concluidos", "Concluídos", "Pedidos finalizados"),
                ),
            )
        )

        filtros = QGroupBox("Filtros")
        filtros.setObjectName("configGroup")
        filtros_layout = QHBoxLayout(filtros)
        filtros_layout.setContentsMargins(20, 15, 20, 15)
        filtros_layout.addWidget(QLabel("Data Inicio:"))
        self.ped_data_inicio = self._create_date_edit(QDate.currentDate().addMonths(-1))
        filtros_layout.addWidget(self.ped_data_inicio)
        filtros_layout.addWidget(QLabel("Data Fim:"))
        self.ped_data_fim = self._create_date_edit(QDate.currentDate())
        filtros_layout.addWidget(self.ped_data_fim)
        filtros_layout.addWidget(QLabel("Status:"))
        self.ped_status = QComboBox()
        self.ped_status.addItems(["Todos", "Pendente", "Aprovado", "Concluído", "Cancelado"])
        self.ped_status.setObjectName("configCombo")
        filtros_layout.addWidget(self.ped_status)
        filtros_layout.addWidget(QLabel("Empresa:"))
        self.ped_empresa = QComboBox()
        self.ped_empresa.addItems(["Todas"])
        self.ped_empresa.setObjectName("configCombo")
        filtros_layout.addWidget(self.ped_empresa)
        filtros_layout.addStretch()
        layout.addWidget(filtros)

        ped_search_card, self.ped_search = self._create_search_card(
            "Pesquisar por material, solicitante, empresa ou link...",
            self._render_pedidos,
        )
        layout.addWidget(ped_search_card)

        self.btn_carregar_pedidos = QPushButton("Atualizar dados")
        self.btn_carregar_pedidos.setObjectName("btnPrimary")
        self.btn_carregar_pedidos.clicked.connect(self.atualizar_pedidos)
        self.btn_exportar_excel_pedidos = QPushButton("Exportar Excel")
        self.btn_exportar_excel_pedidos.setObjectName("btnSecondary")
        self.btn_exportar_excel_pedidos.clicked.connect(lambda: self.exportar_excel("pedidos"))
        self.btn_exportar_pdf_pedidos = QPushButton("Exportar PDF")
        self.btn_exportar_pdf_pedidos.setObjectName("btnSecondary")
        self.btn_exportar_pdf_pedidos.clicked.connect(lambda: self.exportar_pdf("pedidos"))
        layout.addWidget(
            self._create_actions_card(
                self.btn_carregar_pedidos,
                (self.btn_exportar_excel_pedidos, self.btn_exportar_pdf_pedidos),
            )
        )

        self.ped_tabela = self._build_table(
            ["ID", "Material", "Qtd", "Solicitante", "Empresa", "Data Solic.", "Data Conclusão", "Status"],
            stretch_columns=(1,),
            minimum_widths={0: 72, 1: 220, 2: 90, 3: 170, 4: 170, 5: 135, 6: 150, 7: 120},
        )
        self.ped_tabela.horizontalHeader().sectionResized.connect(
            lambda *_args, report_key="pedidos": self._save_table_widths(report_key)
        )
        self.ped_tabela.itemSelectionChanged.connect(lambda: self._update_detail_from_selection("pedidos"))
        ped_content = QHBoxLayout()
        ped_content.setContentsMargins(0, 0, 0, 0)
        ped_content.setSpacing(16)
        ped_content.addWidget(
            self._create_table_panel(
                "pedidos",
                self.ped_tabela,
                "Nenhum pedido encontrado",
                "Ajuste a busca ou revise o periodo, status e empresa para localizar pedidos.",
            ),
            1,
        )
        ped_content.addWidget(
            self._create_detail_panel("pedidos", "Detalhes do pedido", "Selecione um pedido para ver o contexto completo."),
            0,
        )
        layout.addLayout(ped_content)
        return widget

    def create_tab_demandas(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(15)
        layout.addWidget(
            self._create_summary_strip(
                "demandas",
                (
                    ("total", "Demandas", "Chamados no periodo"),
                    ("abertas", "Abertas", "Aguardando tratativa"),
                    ("concluidas", "Concluídas", "Demandas resolvidas"),
                ),
            )
        )

        filtros = QGroupBox("Filtros")
        filtros.setObjectName("configGroup")
        filtros_layout = QHBoxLayout(filtros)
        filtros_layout.setContentsMargins(20, 15, 20, 15)
        filtros_layout.addWidget(QLabel("Data Inicio:"))
        self.dem_data_inicio = self._create_date_edit(QDate.currentDate().addMonths(-1))
        filtros_layout.addWidget(self.dem_data_inicio)
        filtros_layout.addWidget(QLabel("Data Fim:"))
        self.dem_data_fim = self._create_date_edit(QDate.currentDate())
        filtros_layout.addWidget(self.dem_data_fim)
        filtros_layout.addWidget(QLabel("Status:"))
        self.dem_status = QComboBox()
        self.dem_status.addItems(["Todos", "Aberto", "Em Andamento", "Concluído", "Cancelado"])
        self.dem_status.setObjectName("configCombo")
        filtros_layout.addWidget(self.dem_status)
        filtros_layout.addWidget(QLabel("Prioridade:"))
        self.dem_prioridade = QComboBox()
        self.dem_prioridade.addItems(["Todas", "Alta", "Media", "Baixa"])
        self.dem_prioridade.setObjectName("configCombo")
        filtros_layout.addWidget(self.dem_prioridade)
        filtros_layout.addWidget(QLabel("Empresa:"))
        self.dem_empresa = QComboBox()
        self.dem_empresa.addItems(["Todas"])
        self.dem_empresa.setObjectName("configCombo")
        filtros_layout.addWidget(self.dem_empresa)
        filtros_layout.addStretch()
        layout.addWidget(filtros)

        dem_search_card, self.dem_search = self._create_search_card(
            "Pesquisar por título, solicitante, responsável ou descrição...",
            self._render_demandas,
        )
        layout.addWidget(dem_search_card)

        self.btn_carregar_demandas = QPushButton("Atualizar dados")
        self.btn_carregar_demandas.setObjectName("btnPrimary")
        self.btn_carregar_demandas.clicked.connect(self.atualizar_demandas)
        self.btn_exportar_excel_demandas = QPushButton("Exportar Excel")
        self.btn_exportar_excel_demandas.setObjectName("btnSecondary")
        self.btn_exportar_excel_demandas.clicked.connect(lambda: self.exportar_excel("demandas"))
        self.btn_exportar_pdf_demandas = QPushButton("Exportar PDF")
        self.btn_exportar_pdf_demandas.setObjectName("btnSecondary")
        self.btn_exportar_pdf_demandas.clicked.connect(lambda: self.exportar_pdf("demandas"))
        layout.addWidget(
            self._create_actions_card(
                self.btn_carregar_demandas,
                (self.btn_exportar_excel_demandas, self.btn_exportar_pdf_demandas),
            )
        )

        self.dem_tabela = self._build_table(
            ["ID", "Título", "Solicitante", "Prioridade", "Status", "Data Abertura", "Responsável"],
            stretch_columns=(1,),
            minimum_widths={0: 72, 1: 240, 2: 160, 3: 130, 4: 130, 5: 155, 6: 170},
        )
        self.dem_tabela.horizontalHeader().sectionResized.connect(
            lambda *_args, report_key="demandas": self._save_table_widths(report_key)
        )
        self.dem_tabela.itemSelectionChanged.connect(lambda: self._update_detail_from_selection("demandas"))
        dem_content = QHBoxLayout()
        dem_content.setContentsMargins(0, 0, 0, 0)
        dem_content.setSpacing(16)
        dem_content.addWidget(
            self._create_table_panel(
                "demandas",
                self.dem_tabela,
                "Nenhuma demanda encontrada",
                "Ajuste a busca ou revise o periodo, prioridade, status e empresa para encontrar demandas.",
            ),
            1,
        )
        dem_content.addWidget(
            self._create_detail_panel("demandas", "Detalhes da demanda", "Selecione uma demanda para ver o resumo completo."),
            0,
        )
        layout.addLayout(dem_content)
        return widget

    def _create_date_edit(self, value):
        widget = QDateEdit()
        widget.setDate(value)
        widget.setCalendarPopup(True)
        widget.setStyleSheet(CALENDAR_STYLE)
        return widget

    def carregar_empresas_movimentacoes(self):
        self._load_empresas_combo(self.mov_empresa)

    def carregar_empresas_pedidos(self):
        self._load_empresas_combo(self.ped_empresa)

    def carregar_empresas_demandas(self):
        self._load_empresas_combo(self.dem_empresa)

    def _load_empresas_combo(self, combo):
        try:
            empresas = api_client.get_empresas()
            combo.clear()
            combo.addItem("Todas")
            for empresa in empresas:
                if empresa and str(empresa).strip():
                    combo.addItem(str(empresa))
        except Exception as e:
            print(f"Erro ao carregar empresas: {e}")

    def carregar_categorias(self):
        try:
            categorias = api_client.listar_categorias()
            self.est_categoria.clear()
            self.est_categoria.addItem("Todas")
            for categoria in categorias:
                self.est_categoria.addItem(categoria)
            self._load_empresas_combo(self.est_empresa)
        except Exception as e:
            print(f"Erro ao carregar categorias: {e}")

    def _date_edit_to_date(self, widget):
        return widget.date().toPython()

    def _parse_datetime(self, value):
        if not value:
            return None
        raw = str(value).strip()
        if not raw:
            return None
        cleaned = raw.replace("T", " ")
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(cleaned[: len(fmt)], fmt)
            except Exception:
                continue
        return None

    def _passes_period(self, raw_value, start_date, end_date):
        parsed = self._parse_datetime(raw_value)
        if parsed is None:
            return True
        current = parsed.date()
        return start_date <= current <= end_date

    def carregar_movimentacoes(self):
        try:
            self.atualizar_movimentacoes()
        except Exception as e:
            print(f"Erro ao carregar movimentacoes: {e}")
            QMessageBox.warning(self, "Erro", f"Erro ao carregar movimentacoes: {e}")

    def atualizar_movimentacoes(self):
        try:
            movimentacoes = api_client.listar_movimentacoes()
            start_date = self._date_edit_to_date(self.mov_data_inicio)
            end_date = self._date_edit_to_date(self.mov_data_fim)
            tipo = filter_value(self.mov_tipo.currentText())
            empresa = None if is_all_option(self.mov_empresa.currentText()) else self.mov_empresa.currentText()

            filtradas = []
            for item in movimentacoes:
                if not self._passes_period(item.get("data_hora"), start_date, end_date):
                    continue
                if not is_all_option(tipo) and str(item.get("tipo", "")).lower() != tipo:
                    continue
                if empresa and item.get("empresa") != empresa:
                    continue
                filtradas.append(item)

            filtradas.sort(key=lambda x: x.get("data_hora", ""), reverse=True)
            self._report_sources["movimentacoes"] = filtradas
            self._render_movimentacoes()
        except Exception as e:
            print(f"Erro ao atualizar movimentacoes: {e}")

    def _render_movimentacoes(self):
        filtradas = self._apply_text_search(
            self._report_sources["movimentacoes"],
            self.mov_search.text() if hasattr(self, "mov_search") else "",
            ("id", "material_nome", "tipo", "empresa", "destinatario", "usuario_nome", "observacao", "data_hora"),
        )
        self.atualizar_tabela_movimentacoes(filtradas)

    def atualizar_tabela_movimentacoes(self, movimentacoes):
        self._report_visible["movimentacoes"] = list(movimentacoes)
        sorting_enabled = self.mov_tabela.isSortingEnabled()
        self.mov_tabela.setSortingEnabled(False)
        self.mov_tabela.clearContents()
        self.mov_tabela.setRowCount(len(movimentacoes))
        entradas = 0
        saidas = 0

        for row, mov in enumerate(movimentacoes):
            self.mov_tabela.setItem(row, 0, QTableWidgetItem(str(mov.get("id", ""))))
            self.mov_tabela.setItem(row, 1, QTableWidgetItem(mov.get("material_nome", "-")))

            tipo = str(mov.get("tipo", "") or "")
            tipo_item = QTableWidgetItem(tipo.upper())
            if tipo == "entrada":
                entradas += 1
                tipo_item.setForeground(QColor(42, 157, 143))
            else:
                if tipo:
                    saidas += 1
                tipo_item.setForeground(QColor(231, 111, 81))
            self.mov_tabela.setItem(row, 2, tipo_item)

            self.mov_tabela.setItem(row, 3, QTableWidgetItem(str(mov.get("quantidade", 0))))
            self.mov_tabela.setItem(row, 4, QTableWidgetItem(mov.get("empresa", "-")))
            self.mov_tabela.setItem(row, 5, QTableWidgetItem(mov.get("destinatario", "-")))

            data = str(mov.get("data_hora", "") or "")
            if data:
                data = data[:16].replace("T", " ")
            self.mov_tabela.setItem(row, 6, QTableWidgetItem(data))
            self.mov_tabela.setItem(row, 7, QTableWidgetItem(mov.get("usuario_nome", "-")))

            obs = mov.get("observacao")
            self.mov_tabela.setItem(row, 8, QTableWidgetItem("-" if obs is None else str(obs)[:50]))

        if sorting_enabled:
            self.mov_tabela.setSortingEnabled(True)
        refresh_data_table_layout(self.mov_tabela)
        self._apply_saved_table_widths("movimentacoes")
        self._set_summary_value("movimentacoes", "total", len(movimentacoes))
        self._set_summary_value("movimentacoes", "entradas", entradas)
        self._set_summary_value("movimentacoes", "saidas", saidas)
        self._set_report_feedback("movimentacoes", len(movimentacoes))
        self._set_table_empty_state("movimentacoes", bool(movimentacoes))
        if movimentacoes:
            self.mov_tabela.selectRow(0)
            self._update_detail_from_selection("movimentacoes")
        else:
            self._set_detail_empty("movimentacoes")

    def carregar_estoque(self):
        try:
            self.atualizar_estoque()
        except Exception as e:
            print(f"Erro ao carregar estoque: {e}")
            QMessageBox.warning(self, "Erro", f"Erro ao carregar estoque: {e}")

    def atualizar_estoque(self):
        try:
            categoria = None if is_all_option(self.est_categoria.currentText()) else self.est_categoria.currentText()
            empresa = None if is_all_option(self.est_empresa.currentText()) else self.est_empresa.currentText()
            status = "" if is_all_option(self.est_status.currentText()) else filter_value(self.est_status.currentText())
            materiais = api_client.listar_materiais(categoria=categoria, empresa=empresa, status=status)
            materiais.sort(key=lambda x: str(x.get("nome", "")).lower())
            self._report_sources["estoque"] = materiais
            self._render_estoque()
        except Exception as e:
            print(f"Erro ao atualizar estoque: {e}")

    def _render_estoque(self):
        materiais = self._apply_text_search(
            self._report_sources["estoque"],
            self.est_search.text() if hasattr(self, "est_search") else "",
            ("id", "nome", "descricao", "categoria", "empresa", "status", "quantidade"),
        )
        self.atualizar_tabela_estoque(materiais)

    def atualizar_tabela_estoque(self, materiais):
        self._report_visible["estoque"] = list(materiais)
        sorting_enabled = self.est_tabela.isSortingEnabled()
        self.est_tabela.setSortingEnabled(False)
        self.est_tabela.clearContents()
        self.est_tabela.setRowCount(len(materiais))
        ativos = 0
        criticos = 0

        for row, mat in enumerate(materiais):
            quantidade = int(mat.get("quantidade", 0) or 0)
            status = str(mat.get("status", "ativo") or "ativo")
            if status == "ativo":
                ativos += 1
            if quantidade <= 2:
                criticos += 1

            self.est_tabela.setItem(row, 0, QTableWidgetItem(str(mat.get("id", ""))))
            self.est_tabela.setItem(row, 1, QTableWidgetItem(mat.get("nome", "")))
            self.est_tabela.setItem(row, 2, QTableWidgetItem(str(mat.get("descricao", "") or "")[:60]))
            self.est_tabela.setItem(row, 3, QTableWidgetItem(str(quantidade)))
            self.est_tabela.setItem(row, 4, QTableWidgetItem(mat.get("categoria", "-")))
            self.est_tabela.setItem(row, 5, QTableWidgetItem(mat.get("empresa", "-")))

            status_item = QTableWidgetItem(status.upper())
            status_item.setForeground(QColor(42, 157, 143) if status == "ativo" else QColor(231, 111, 81))
            self.est_tabela.setItem(row, 6, status_item)

        if sorting_enabled:
            self.est_tabela.setSortingEnabled(True)
        refresh_data_table_layout(self.est_tabela)
        self._apply_saved_table_widths("estoque")
        self._set_summary_value("estoque", "total", len(materiais))
        self._set_summary_value("estoque", "ativos", ativos)
        self._set_summary_value("estoque", "criticos", criticos)
        self._set_report_feedback("estoque", len(materiais))
        self._set_table_empty_state("estoque", bool(materiais))
        if materiais:
            self.est_tabela.selectRow(0)
            self._update_detail_from_selection("estoque")
        else:
            self._set_detail_empty("estoque")

    def carregar_pedidos(self):
        try:
            self.atualizar_pedidos()
        except Exception as e:
            print(f"Erro ao carregar pedidos: {e}")
            QMessageBox.warning(self, "Erro", f"Erro ao carregar pedidos: {e}")

    def atualizar_pedidos(self):
        try:
            status = None if is_all_option(self.ped_status.currentText()) else filter_value(self.ped_status.currentText())
            empresa = None if is_all_option(self.ped_empresa.currentText()) else self.ped_empresa.currentText()
            pedidos = api_client.listar_pedidos(status=status, empresa=empresa)

            start_date = self._date_edit_to_date(self.ped_data_inicio)
            end_date = self._date_edit_to_date(self.ped_data_fim)
            filtrados = [p for p in pedidos if self._passes_period(p.get("data_solicitacao"), start_date, end_date)]
            filtrados.sort(key=lambda x: x.get("data_solicitacao", ""), reverse=True)
            self._report_sources["pedidos"] = filtrados
            self._render_pedidos()
        except Exception as e:
            print(f"Erro ao atualizar pedidos: {e}")

    def _render_pedidos(self):
        pedidos = self._apply_text_search(
            self._report_sources["pedidos"],
            self.ped_search.text() if hasattr(self, "ped_search") else "",
            ("id", "material_nome", "solicitante", "empresa", "departamento", "status", "observacao", "link_compra"),
        )
        self.atualizar_tabela_pedidos(pedidos)

    def atualizar_tabela_pedidos(self, pedidos):
        self._report_visible["pedidos"] = list(pedidos)
        sorting_enabled = self.ped_tabela.isSortingEnabled()
        self.ped_tabela.setSortingEnabled(False)
        self.ped_tabela.clearContents()
        self.ped_tabela.setRowCount(len(pedidos))
        pendentes = 0
        concluidos = 0
        status_cores = {
            "pendente": QColor(244, 162, 97),
            "aprovado": QColor(42, 157, 143),
            "concluido": QColor(44, 125, 160),
            "cancelado": QColor(231, 111, 81),
        }

        for row, ped in enumerate(pedidos):
            status = str(ped.get("status", "pendente") or "pendente")
            if status == "pendente":
                pendentes += 1
            if status == "concluido":
                concluidos += 1

            self.ped_tabela.setItem(row, 0, QTableWidgetItem(str(ped.get("id", ""))))
            self.ped_tabela.setItem(row, 1, QTableWidgetItem(ped.get("material_nome", "-")))
            self.ped_tabela.setItem(row, 2, QTableWidgetItem(str(ped.get("quantidade", 0))))
            self.ped_tabela.setItem(row, 3, QTableWidgetItem(ped.get("solicitante", "-")))
            self.ped_tabela.setItem(row, 4, QTableWidgetItem(ped.get("empresa", "-")))
            self.ped_tabela.setItem(row, 5, QTableWidgetItem(ped.get("data_solicitacao", "-")))
            self.ped_tabela.setItem(row, 6, QTableWidgetItem(ped.get("data_conclusao", "-") or "-"))

            status_item = QTableWidgetItem(status.upper())
            status_item.setForeground(status_cores.get(status, QColor(0, 0, 0)))
            self.ped_tabela.setItem(row, 7, status_item)

        if sorting_enabled:
            self.ped_tabela.setSortingEnabled(True)
        refresh_data_table_layout(self.ped_tabela)
        self._apply_saved_table_widths("pedidos")
        self._set_summary_value("pedidos", "total", len(pedidos))
        self._set_summary_value("pedidos", "pendentes", pendentes)
        self._set_summary_value("pedidos", "concluidos", concluidos)
        self._set_report_feedback("pedidos", len(pedidos))
        self._set_table_empty_state("pedidos", bool(pedidos))
        if pedidos:
            self.ped_tabela.selectRow(0)
            self._update_detail_from_selection("pedidos")
        else:
            self._set_detail_empty("pedidos")

    def carregar_demandas(self):
        try:
            self.atualizar_demandas()
        except Exception as e:
            print(f"Erro ao carregar demandas: {e}")
            QMessageBox.warning(self, "Erro", f"Erro ao carregar demandas: {e}")

    def atualizar_demandas(self):
        try:
            status = None if is_all_option(self.dem_status.currentText()) else filter_value(self.dem_status.currentText())
            prioridade = None if is_all_option(self.dem_prioridade.currentText()) else filter_value(self.dem_prioridade.currentText())
            empresa = None if is_all_option(self.dem_empresa.currentText()) else self.dem_empresa.currentText()
            demandas = api_client.listar_demandas(status=status, prioridade=prioridade, empresa=empresa)

            start_date = self._date_edit_to_date(self.dem_data_inicio)
            end_date = self._date_edit_to_date(self.dem_data_fim)
            filtradas = [d for d in demandas if self._passes_period(d.get("data_abertura"), start_date, end_date)]
            filtradas.sort(key=lambda x: x.get("data_abertura", ""), reverse=True)
            self._report_sources["demandas"] = filtradas
            self._render_demandas()
        except Exception as e:
            print(f"Erro ao atualizar demandas: {e}")

    def _render_demandas(self):
        demandas = self._apply_text_search(
            self._report_sources["demandas"],
            self.dem_search.text() if hasattr(self, "dem_search") else "",
            ("id", "titulo", "descricao", "solicitante", "responsavel", "empresa", "status", "prioridade", "observacao"),
        )
        self.atualizar_tabela_demandas(demandas)

    def atualizar_tabela_demandas(self, demandas):
        self._report_visible["demandas"] = list(demandas)
        sorting_enabled = self.dem_tabela.isSortingEnabled()
        self.dem_tabela.setSortingEnabled(False)
        self.dem_tabela.clearContents()
        self.dem_tabela.setRowCount(len(demandas))
        abertas = 0
        concluidas = 0
        prioridade_cores = {
            "alta": QColor(231, 111, 81),
            "media": QColor(244, 162, 97),
            "baixa": QColor(42, 157, 143),
        }

        for row, dem in enumerate(demandas):
            status = str(dem.get("status", "aberto") or "aberto")
            prioridade = str(dem.get("prioridade", "media") or "media")
            if status in {"aberto", "em_andamento"}:
                abertas += 1
            if status == "concluido":
                concluidas += 1

            self.dem_tabela.setItem(row, 0, QTableWidgetItem(str(dem.get("id", ""))))
            self.dem_tabela.setItem(row, 1, QTableWidgetItem(str(dem.get("titulo", "") or "")[:60]))
            self.dem_tabela.setItem(row, 2, QTableWidgetItem(dem.get("solicitante", "-")))

            prioridade_item = QTableWidgetItem(prioridade.upper())
            prioridade_item.setForeground(prioridade_cores.get(prioridade, QColor(0, 0, 0)))
            self.dem_tabela.setItem(row, 3, prioridade_item)

            status_item = QTableWidgetItem(status.upper())
            self.dem_tabela.setItem(row, 4, status_item)

            data = str(dem.get("data_abertura", "") or "")
            if data:
                data = data[:10]
            self.dem_tabela.setItem(row, 5, QTableWidgetItem(data))
            self.dem_tabela.setItem(row, 6, QTableWidgetItem(dem.get("responsavel", "-")))

        if sorting_enabled:
            self.dem_tabela.setSortingEnabled(True)
        refresh_data_table_layout(self.dem_tabela)
        self._apply_saved_table_widths("demandas")
        self._set_summary_value("demandas", "total", len(demandas))
        self._set_summary_value("demandas", "abertas", abertas)
        self._set_summary_value("demandas", "concluidas", concluidas)
        self._set_report_feedback("demandas", len(demandas))
        self._set_table_empty_state("demandas", bool(demandas))
        if demandas:
            self.dem_tabela.selectRow(0)
            self._update_detail_from_selection("demandas")
        else:
            self._set_detail_empty("demandas")

    def _table_for_tipo(self, tipo):
        attr_map = {
            "movimentacoes": "mov_tabela",
            "estoque": "est_tabela",
            "pedidos": "ped_tabela",
            "demandas": "dem_tabela",
        }
        attr_name = attr_map.get(tipo)
        if not attr_name:
            return None
        return getattr(self, attr_name, None)

    def _table_to_rows(self, table):
        headers = []
        for column in range(table.columnCount()):
            item = table.horizontalHeaderItem(column)
            headers.append(item.text() if item else f"Coluna {column + 1}")

        rows = []
        for row in range(table.rowCount()):
            row_values = []
            for column in range(table.columnCount()):
                cell_widget = table.cellWidget(row, column)
                if cell_widget is not None:
                    labels = cell_widget.findChildren(QLabel)
                    if labels:
                        row_values.append(labels[0].text())
                        continue
                item = table.item(row, column)
                row_values.append(item.text() if item else "")
            rows.append(row_values)
        return headers, rows

    def _report_label(self, tipo):
        labels = {
            "movimentacoes": "movimentacoes",
            "estoque": "estoque",
            "pedidos": "pedidos",
            "demandas": "demandas",
        }
        return labels.get(tipo, tipo)

    def _sanitize_filename_part(self, value):
        text = str(value or "").strip()
        if not text:
            return ""
        text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
        text = text.lower()
        text = re.sub(r"[^a-z0-9]+", "_", text)
        return text.strip("_")

    def _current_export_filters(self, tipo):
        filters = []

        if tipo == "movimentacoes":
            filters.append(("inicio", self.mov_data_inicio.date().toString("yyyy-MM-dd")))
            filters.append(("fim", self.mov_data_fim.date().toString("yyyy-MM-dd")))
            if not is_all_option(self.mov_tipo.currentText()):
                filters.append(("tipo", self.mov_tipo.currentText()))
            if not is_all_option(self.mov_empresa.currentText()):
                filters.append(("empresa", self.mov_empresa.currentText()))
        elif tipo == "estoque":
            if not is_all_option(self.est_categoria.currentText()):
                filters.append(("categoria", self.est_categoria.currentText()))
            if not is_all_option(self.est_empresa.currentText()):
                filters.append(("empresa", self.est_empresa.currentText()))
            if not is_all_option(self.est_status.currentText()):
                filters.append(("status", self.est_status.currentText()))
        elif tipo == "pedidos":
            filters.append(("inicio", self.ped_data_inicio.date().toString("yyyy-MM-dd")))
            filters.append(("fim", self.ped_data_fim.date().toString("yyyy-MM-dd")))
            if not is_all_option(self.ped_status.currentText()):
                filters.append(("status", self.ped_status.currentText()))
            if not is_all_option(self.ped_empresa.currentText()):
                filters.append(("empresa", self.ped_empresa.currentText()))
        elif tipo == "demandas":
            filters.append(("inicio", self.dem_data_inicio.date().toString("yyyy-MM-dd")))
            filters.append(("fim", self.dem_data_fim.date().toString("yyyy-MM-dd")))
            if not is_all_option(self.dem_status.currentText()):
                filters.append(("status", self.dem_status.currentText()))
            if not is_all_option(self.dem_prioridade.currentText()):
                filters.append(("prioridade", self.dem_prioridade.currentText()))
            if not is_all_option(self.dem_empresa.currentText()):
                filters.append(("empresa", self.dem_empresa.currentText()))

        return filters

    def _build_export_basename(self, tipo):
        parts = [f"relatorio_{self._sanitize_filename_part(self._report_label(tipo))}"]
        for key, value in self._current_export_filters(tipo):
            sanitized = self._sanitize_filename_part(value)
            if sanitized:
                parts.append(f"{key}_{sanitized}")
        parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
        return "_".join(parts)

    def _export_context_text(self, tipo, row_count):
        filters = self._current_export_filters(tipo)
        filter_text = ", ".join(f"{label}: {value}" for label, value in filters) if filters else "sem filtros adicionais"
        return (
            f"Relatorio exportado com sucesso.\n\n"
            f"Registros exportados: {row_count}\n"
            f"Filtros: {filter_text}"
        )

    def exportar_excel(self, tipo):
        if not self._pode("relatorios.export"):
            self._avisar_sem_permissao("relatorios.export")
            return

        table = self._table_for_tipo(tipo)
        if table is None:
            return

        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Salvar Excel",
                f"{self._build_export_basename(tipo)}.xlsx",
                "Excel Files (*.xlsx)",
            )
            if not file_path:
                return

            headers, rows = self._table_to_rows(table)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"Relatorio {tipo.capitalize()}"

            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
            title_cell = sheet.cell(row=1, column=1)
            title_cell.value = f"Relatorio de {tipo.capitalize()}"
            title_cell.font = Font(name="Arial", size=16, bold=True)
            title_cell.alignment = Alignment(horizontal="center")

            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(headers)))
            date_cell = sheet.cell(row=2, column=1)
            date_cell.value = f"Emitido em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            date_cell.alignment = Alignment(horizontal="center")

            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            for column, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=column, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row_index, row_values in enumerate(rows, start=5):
                for column, value in enumerate(row_values, start=1):
                    cell = sheet.cell(row=row_index, column=column, value=value)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = border

            for index, header in enumerate(headers, start=1):
                max_width = max(len(str(header)), *(len(str(row[index - 1])) for row in rows), 12)
                sheet.column_dimensions[get_column_letter(index)].width = min(max_width + 2, 40)

            workbook.save(file_path)
            QMessageBox.information(
                self,
                "Sucesso",
                f"{self._export_context_text(tipo, len(rows))}\nArquivo: {file_path}",
            )
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar Excel: {e}")

    def exportar_pdf(self, tipo):
        if not self._pode("relatorios.export"):
            self._avisar_sem_permissao("relatorios.export")
            return

        table_widget = self._table_for_tipo(tipo)
        if table_widget is None:
            return

        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Salvar PDF",
                f"{self._build_export_basename(tipo)}.pdf",
                "PDF Files (*.pdf)",
            )
            if not file_path:
                return

            headers, rows = self._table_to_rows(table_widget)
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("ReportTitle", parent=styles["Heading1"], fontSize=18, leading=22, spaceAfter=8)
            meta_style = ParagraphStyle("ReportMeta", parent=styles["BodyText"], fontSize=9, textColor=colors.HexColor("#475569"))
            cell_style = ParagraphStyle("ReportCell", parent=styles["BodyText"], fontSize=8, leading=10)

            elements = [
                Paragraph(f"Relatorio de {tipo.capitalize()}", title_style),
                Paragraph(f"Emitido em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", meta_style),
                Spacer(1, 10),
            ]

            data = [[Paragraph(str(header), cell_style) for header in headers]]
            for row_values in rows:
                data.append([Paragraph(str(value), cell_style) for value in row_values])

            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(table)
            doc.build(elements)
            QMessageBox.information(
                self,
                "Sucesso",
                f"{self._export_context_text(tipo, len(rows))}\nArquivo: {file_path}",
            )
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar PDF: {e}")
